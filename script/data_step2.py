# 本脚本是针对中国新冠病毒各省市历史发病数据的清洗工具
# 作者 https://github.com/Avens666  mail: cz_666@qq.com
# 数据源使用 https://github.com/BlankerL/DXY-COVID-19-Data/blob/master/csv/DXYArea.csv
# 输入源数据来自 data_step1.py 的输出文件
# 本脚本基于data_step1.py的输出 计算每天的新增数据，源数据只有每天的累计确诊数据，本脚本通过当天数据减去前一天数据的方式，计算出每天新增数据
# 用户通过修改 inputfile  和  outputfile 定义源数据文件和输出文件

from datetime import timedelta

import pandas
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 输入输出参数，根据需要设置
input_file = "out1.csv"
output_file = "out_2_23.csv"
excel_file = "全国疫情数据分析.xlsx"  # "test.xlsx"
b_export_excel = False

# 显示所有列
pandas.set_option('display.max_columns', None)
# 显示所有行
pandas.set_option('display.max_rows', None)
# 设置value的显示长度为100，默认为50
pandas.set_option('max_colwidth', 200)

# ！！！ 根据需要选择合适的字符集
try:
    dataf = pandas.read_csv(input_file, encoding='UTF-8')
except:
    dataf = pandas.read_csv(input_file, encoding='gb2312')

# 计算增量 根据日期间隔计算

dataf['日期'] = pandas.to_datetime(dataf['日期'], format='%Y-%m-%d')  # 1900 -> 2020

df_t = dataf['日期']
df_date = df_t.drop_duplicates()  # 去重 这个返回Series对象

# dataf['新增确诊'] = dataf['确诊']
dataf.insert(loc=5, column='新增确诊', value=0)
dataf.insert(loc=6, column='新增治愈', value=0)
dataf.insert(loc=7, column='新增死亡', value=0)

# df_date = df_date.sort_values(ascending=False)
min_date = df_date.min()
cur_date = df_date.max()

df_t = pandas.DataFrame()

step = 0
print('处理数据... ')
for index, data in dataf.iterrows():
    data2 = dataf.loc[
            (dataf['省'] == data['省']) & (dataf['市'] == data['市']) & (dataf['日期'] == data['日期'] - timedelta(days=1)),
            :]
    # 判断后一天是否存在
    data3 = dataf.loc[
            (dataf['省'] == data['省']) & (dataf['市'] == data['市']) & (dataf['日期'] == data['日期'] + timedelta(days=1)),
            :]

    if data2.shape[0] > 0:
        dataf.loc[index, '新增确诊'] = data['确诊'] - data2.iloc[0, :]['确诊']
        dataf.loc[index, '新增治愈'] = data['治愈'] - data2.iloc[0, :]['治愈']
        dataf.loc[index, '新增死亡'] = data['死亡'] - data2.iloc[0, :]['死亡']
    else:
        dataf.loc[index, '新增确诊'] = data['确诊']
        dataf.loc[index, '新增治愈'] = data['治愈']
        dataf.loc[index, '新增死亡'] = data['死亡']

    if (data['日期'] != cur_date) & (data3.shape[0] == 0):  # 后一天如果没有数据，加一条冲抵数据
        new = pandas.DataFrame({'省': data['省'],
                                '省确诊': data['省确诊'],
                                '省治愈': data['省治愈'],
                                '省死亡': data['省死亡'],
                                '市': data['市'],
                                '新增确诊': 0 - data['确诊'],
                                '新增治愈': 0 - data['治愈'],
                                '新增死亡': 0 - data['死亡'],
                                '确诊': 0,
                                '治愈': 0,
                                '死亡': 0,
                                '日期': data['日期'] + timedelta(days=1)},
                               pandas.Index(range(1)))
        #            print(new.head())
        df_t = df_t.append(new)
    step += 1
    if step % 2000 == 0:
        print('处理数据条数: ' + str(step))
#    print(data['日期'])  # 输出处理进度
print('处理数据完毕 条数: ' + str(step))
out = dataf.append(df_t, sort=False)

# 日期矫正为前一天，因为当天是发布日期，数据是反映前一天日期
out['日期'] = out['日期'] - timedelta(days=1)

print('输出csv文件 : ' + output_file)
out.to_csv(output_file, encoding="utf_8_sig", index=False)  # 为保证excel打开兼容，输出为UTF8带签名格式


# 导出excel 可选
if b_export_excel:
    writer = pandas.ExcelWriter(excel_file, engine='openpyxl')
    try:
        # 加载指定的excel文件
        print('加载excel文件 : ' + excel_file)
        writer.book = load_workbook(excel_file)

        print('处理数据...')
        sheet_name = '原始数据'
        max_row = writer.book[sheet_name].max_row

        writer.book[sheet_name].delete_rows(2, max_row-1)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    out.to_excel(writer, sheet_name, startrow=1, index=False, header=False)

    print('输出excel文件' + excel_file)
    writer.save()

